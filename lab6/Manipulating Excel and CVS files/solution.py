import openpyxl
import sys
import csv


input_file_name = sys.argv[1]
filename =input_file_name+'.csv'
input_workno = sys.argv[2]

examplefile=open(filename)
exampleReader=csv.reader(examplefile)
exampleData=list(exampleReader)

headers=exampleData[0]

#find column number of work no in headers of input file.
for i in range(len(headers)):
	header=headers[i]
	header=header.lower()
	if 'work no' in header:
		col= i
		break
data=[]
#Extracting all the rows that contains desired 'input_workno' in work no column.
for r in range(len(exampleData)):
	row_data=exampleData[r]
	if(row_data[col]==input_workno):
		data.append(row_data)

#creating workbook object.
wb = openpyxl.Workbook()
#deleting default sheet in workbook.
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

#creating sheets to write every single touples. 
for k in range(len(data)):
	wb.create_sheet(input_workno+'_'+str(k+1))

new_sheets=wb.get_sheet_names()
#Writting each touples of data into different sheets.
for i in range(len(data)):
	sheet=wb.get_sheet_by_name(new_sheets[i])
	for j in range(len(headers)):
		sheet.cell(row=1,column=j+1).value=headers[j]
	row_data=data[i]
	for j in range(len(row_data)):
		if(j==col):
			sheet.cell(row=2,column=j+1).value=row_data[j]+'_'+str(i+1)
		else:
			sheet.cell(row=2,column=j+1).value=row_data[j]

#saving file by the desired name.
wb.save(input_file_name+'_'+input_workno+'.xlsx')
        

	
	
	



	

